import time
import unittest
import os
import xlrd
import logging
import traceback
from datetime import datetime
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options
from faker import Faker
from openpyxl import load_workbook
import locale
import warnings

from AUTO.PageApp.loginPage import LoginPage
from AUTO.PageApp.searchPage import SearchPage
from AUTO.PageApp.quotePage import QuotePage
from DATAENTRY_DATOS_PERSONALES import solicita_credito, solicitud_datos_personales
from QUOTE_PRODUCTO import cotizador_producto
from QUOTE_PLAN import cotizador_plan_financiero
from QUOTE_SEGURO import cotizador_seguro_de_auto
from QUOTE_DETALLE_FINANCIAMIENTO import cotizador_detalle_financiamiento
from QUOTE_DATOS_SOLICITANTE import cotizador_datos_solicitante
from QUOTE_IMPRIMIR import cotizador_imprimir

# Suppress UserWarnings from openpyxl
warnings.simplefilter(action='ignore', category=UserWarning)

# Files
BASE_DIR = 'C:/Automation_Performance_Sast'
LOG_DIR_BASE = os.path.join(BASE_DIR, 'app_log')
DRIVER_PATH = 'C:/Users/jgarcia/PycharmProjects/SNTDProject/drivers/Edge/126/msedgedriver.exe'
EXCEL_FILE_QUOTE = os.path.join(BASE_DIR, 'SNTD_QUOTE_PF_TEST.xls')
EXCEL_FILE_COT = os.path.join(BASE_DIR, 'Cotizador_CAT_2.xlsm')

# Configuración de Log
log_subdir = datetime.now().strftime('%Y-%m-%d')
log_dir = os.path.join(LOG_DIR_BASE, log_subdir)
os.makedirs(log_dir, exist_ok=True)
log_filename = os.path.join(log_dir, 'log.txt')
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Inicializa el método Faker
fake = Faker()


class MyTestCase(unittest.TestCase):
    driver = None

    def __init__(self, *args, **kwargs):
        super(MyTestCase, self).__init__(*args, **kwargs)
        self.final_string = ""
        self.commentgap = ""
        self.commentrsa = ""
        self.commentge = ""
        self.ballon_and_tcm = ''
        self.TblamortizationTCM = ''
        self.namecomplete = ''
        self.productall = ''
        self.lblcar = ''
        self.lblpricelist = ''
        self.tdaccessoriesamount = ''
        self.tdadditionalcoverageamount = ''
        self.tdextendedwarrantyamount = ''
        self.tdautopartstheftinsuranceamount = ''
        self.tdInsuranceGAPAmount = ''
        self.tdTasa = ''
        self.tdPLazo = ''
        self.lblSeguroVd = ''
        self.tdSeguro = ''
        self.replace_divInsuranceRecipes = ''
        self.lblEnganche = ''
        self.tasaC2 = ''
        self.plazoC2 = ''
        self.tdComision = ''
        self.tdDesembolso = ''
        self.tdFinanciar = ''
        self.tdMensualidad = ''
        self.cat = ''
        self.cuotasall = ''
        self.Tblamortization = ''

    @classmethod
    def setUpClass(cls):
        edge_service = EdgeService(DRIVER_PATH)
        edge_options = Options()
        edge_options.add_argument('--ignore-certificate-errors')
        edge_options.add_argument('--ignore-ssl-errors')
        # edge_options.headless = True
        cls.driver = webdriver.Edge(service=edge_service, options=edge_options)
        cls.driver.delete_all_cookies()
        if "data:" not in cls.driver.current_url:
            cls.driver.execute_script("localStorage.clear();")
            cls.driver.execute_script("sessionStorage.clear();")
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_cotizador_valid(self):
        archivo_excel = xlrd.open_workbook(EXCEL_FILE_QUOTE)
        sheet_environment = archivo_excel.sheet_by_index(0)
        fila_environment = {
            "var_environment": sheet_environment.cell_value(1, 0),
            "var_url": sheet_environment.cell_value(1, 1),
            "var_canal": sheet_environment.cell_value(1, 2)
        }
        cot_excel = load_workbook(filename=EXCEL_FILE_COT, data_only=True)
        sheet_cot = cot_excel['Cotizador']
        fila_cot = {
            "precio_vehiculo": sheet_cot.cell(row=5, column=3).value,
            "enganche": sheet_cot.cell(row=6, column=3).value,
            "aforo": sheet_cot.cell(row=7, column=3).value,
            "seguro_danos_1": sheet_cot.cell(row=11, column=3).value,
            "seguro_danos_2": sheet_cot.cell(row=12, column=3).value,
            "seguro_danos_subsecuente": sheet_cot.cell(row=13, column=3).value,
            "suma_subsecuente": sheet_cot.cell(row=14, column=3).value,
            "seguro_vida": sheet_cot.cell(row=15, column=3).value,
            "accesorios_garantia_financiado": sheet_cot.cell(row=16, column=3).value,
            "accesorios_garantia_contado": sheet_cot.cell(row=17, column=3).value,
            "garantia_extendida": sheet_cot.cell(row=14, column=4).value,
            "seguro_autoparte": sheet_cot.cell(row=16, column=4).value,
            "seguro_gap_porcentaje": sheet_cot.cell(row=18, column=5).value,
            "seguro_gap_monto": sheet_cot.cell(row=18, column=4).value,
            "accesorios": sheet_cot.cell(row=20, column=4).value,
            "coberturas": sheet_cot.cell(row=22, column=4).value,
            "monto_financiar": sheet_cot.cell(row=18, column=3).value,
            "plazo": sheet_cot.cell(row=19, column=3).value,
            "tasa": sheet_cot.cell(row=20, column=3).value,
            "comision_porcentaje": sheet_cot.cell(row=21, column=3).value,
            "comision_monto": sheet_cot.cell(row=22, column=3).value,
            "desembolso": sheet_cot.cell(row=24, column=3).value,
            "CAT": sheet_cot.cell(row=26, column=3).value,
            "fecha": sheet_cot.cell(row=1, column=9).value,
            "fecha_pago_0": sheet_cot.cell(row=35, column=3).value,
            "interes_0": sheet_cot.cell(row=35, column=6).value,
            "iva_interes_0": sheet_cot.cell(row=35, column=8).value,
            "mensualidad_0": sheet_cot.cell(row=35, column=9).value,
            "fecha_pago_1": sheet_cot.cell(row=36, column=3).value,
            "capital_1": sheet_cot.cell(row=36, column=5).value,
            "interes_1": sheet_cot.cell(row=36, column=6).value,
            "iva_interes_1": sheet_cot.cell(row=36, column=8).value,
            "mensualidad_1": sheet_cot.cell(row=36, column=9).value,
            "fecha_pago_13": sheet_cot.cell(row=48, column=3).value,
            "capital_13": sheet_cot.cell(row=48, column=5).value,
            "interes_13": sheet_cot.cell(row=48, column=6).value,
            "iva_interes_13": sheet_cot.cell(row=48, column=8).value,
            "mensualidad_13": sheet_cot.cell(row=48, column=9).value,
            "fecha_pago_25": sheet_cot.cell(row=60, column=3).value,
            "capital_25": sheet_cot.cell(row=60, column=5).value,
            "interes_25": sheet_cot.cell(row=60, column=6).value,
            "iva_interes_25": sheet_cot.cell(row=60, column=8).value,
            "mensualidad_25": sheet_cot.cell(row=60, column=9).value,
            "fecha_pago_37": sheet_cot.cell(row=72, column=3).value,
            "capital_37": sheet_cot.cell(row=72, column=5).value,
            "interes_37": sheet_cot.cell(row=72, column=6).value,
            "iva_interes_37": sheet_cot.cell(row=72, column=8).value,
            "mensualidad_37": sheet_cot.cell(row=72, column=9).value,
        }
        locale.setlocale(locale.LC_ALL, '')
        quote_precio_vehiculo = locale.currency(fila_cot["precio_vehiculo"], grouping=True)
        quote_enganche = locale.currency(fila_cot["enganche"], grouping=True)
        quote_aforo = locale.currency(fila_cot["aforo"], grouping=True)
        quote_seguro_danos_1 = locale.currency(fila_cot["seguro_danos_1"], grouping=True)
        quote_seguro_danos_2 = locale.currency(fila_cot["seguro_danos_2"], grouping=True)
        quote_seguro_danos_subsecuente = locale.currency(fila_cot["seguro_danos_subsecuente"], grouping=True)
        quote_suma_subsecuente = locale.currency(fila_cot["suma_subsecuente"], grouping=True)
        quote_seguro_vida = locale.currency(fila_cot["seguro_vida"], grouping=True)
        quote_ge_acc_financiado = locale.currency(fila_cot["accesorios_garantia_financiado"], grouping=True)
        quote_ge_acc_contado = locale.currency(fila_cot["accesorios_garantia_contado"], grouping=True)
        quote_garantia_extendida = locale.currency(fila_cot["garantia_extendida"], grouping=True)
        quote_seguro_autoparte = locale.currency(fila_cot["seguro_autoparte"], grouping=True)
        quote_seguro_gap_porcentaje = format(fila_cot["seguro_gap_porcentaje"], ".1%")
        quote_seguro_gap_monto = locale.currency(fila_cot["seguro_gap_monto"], grouping=True)
        quote_accesorios_financiados = locale.currency(fila_cot["accesorios"], grouping=True)
        quote_coberturas = locale.currency(fila_cot["coberturas"], grouping=True)
        quote_monto_financiar = locale.currency(fila_cot["monto_financiar"], grouping=True)
        quote_plazo = fila_cot["plazo"]
        quote_tasa = format(fila_cot["tasa"], ".2%")
        quote_comision_porcentaje = format(fila_cot["comision_porcentaje"], ".2%")
        quote_comision_monto = locale.currency(fila_cot["comision_monto"], grouping=True)
        quote_desembolso = locale.currency(fila_cot["desembolso"], grouping=True)
        quote_cat = format(fila_cot["CAT"], ".1%")
        quote_fecha = fila_cot["fecha"].strftime("%d-%m-%Y")
        quote_fecha_pago_0 = fila_cot["fecha_pago_0"].strftime("%d-%m-%Y")
        info_0 = (
            f"Fecha de pago: {quote_fecha_pago_0}",
            f"Interés: {locale.currency(fila_cot['interes_0'], grouping=True)}",
            f"IVA de interés: {locale.currency(fila_cot['iva_interes_0'], grouping=True)}",
            f"Pago Mensual: {locale.currency(fila_cot['mensualidad_0'], grouping=True)}"
        )
        _info_0 = ', '.join(info_0)
        quote_fecha_pago_1 = fila_cot["fecha_pago_1"].strftime("%d-%m-%Y")
        info_1 = (
            f"Fecha de pago: {quote_fecha_pago_1}",
            f"Capital: {locale.currency(fila_cot['capital_1'], grouping=True)}",
            f"Interés: {locale.currency(fila_cot['interes_1'], grouping=True)}",
            f"IVA de interés: {locale.currency(fila_cot['iva_interes_1'], grouping=True)}",
            f"Pago Mensual: {locale.currency(fila_cot['mensualidad_1'], grouping=True)}"
        )
        _info_1 = ', '.join(info_1)
        if fila_cot["fecha_pago_13"] is not None:
            quote_fecha_pago_13 = fila_cot["fecha_pago_13"].strftime("%d-%m-%Y")
            capital_13 = locale.currency(fila_cot['capital_13'], grouping=True)
            interes_13 = locale.currency(fila_cot['interes_13'], grouping=True)
            iva_interes_13 = locale.currency(fila_cot['iva_interes_13'], grouping=True)
            mensualidad_13 = locale.currency(fila_cot['mensualidad_13'], grouping=True)
        else:
            quote_fecha_pago_13 = 'NA'
            capital_13 = 'NA'
            interes_13 = 'NA'
            iva_interes_13 = 'NA'
            mensualidad_13 = 'NA'

        info_13 = (
            f"Fecha de pago: {quote_fecha_pago_13}",
            f"Capital: {capital_13}",
            f"Interés: {interes_13}",
            f"IVA de interés: {iva_interes_13}",
            f"Pago Mensual: {mensualidad_13}"
        )
        _info_13 = ', '.join(info_13)

        if fila_cot["fecha_pago_25"] is not None:
            quote_fecha_pago_25 = fila_cot["fecha_pago_25"].strftime("%d-%m-%Y")
            capital_25 = locale.currency(fila_cot['capital_25'], grouping=True)
            interes_25 = locale.currency(fila_cot['interes_25'], grouping=True)
            iva_interes_25 = locale.currency(fila_cot['iva_interes_25'], grouping=True)
            mensualidad_25 = locale.currency(fila_cot['mensualidad_25'], grouping=True)
        else:
            quote_fecha_pago_25 = 'NA'
            capital_25 = 'NA'
            interes_25 = 'NA'
            iva_interes_25 = 'NA'
            mensualidad_25 = 'NA'

        info_25 = (
            f"Fecha de pago: {quote_fecha_pago_25}",
            f"Capital: {capital_25}",
            f"Interés: {interes_25}",
            f"IVA de interés: {iva_interes_25}",
            f"Pago Mensual: {mensualidad_25}"
        )
        _info_25 = ', '.join(info_25)

        # print("\033[1;32m_" * 10 + " DATOS DEL SIMULADOR DEL EXCEL ", "_\033[0m" * 10)
        # print(
        #     "Precio del Vehículo: ", quote_precio_vehiculo,
        #     "\nEnganche: ", quote_enganche,
        #     "\nAforo: ", quote_aforo,
        #     "\nSeguro de Daños Año 1: ", quote_seguro_danos_1,
        #     "\nSeguro de Daños Año 2: ", quote_seguro_danos_2,
        #     "\nSeguro de Daños Subsecuentes: ", quote_seguro_danos_subsecuente,
        #     "\nSuma subsecuentes : ", quote_suma_subsecuente,
        #     "\nSeguro de Vida y Desempleo: ", quote_seguro_vida,
        #     "\nGarantía y/o Accesorios Financiados: ", quote_ge_acc_financiado,
        #     "\nGarantía y/o Accesorios Contado: ", quote_ge_acc_contado,
        #     "\nGarantía Extendida: ", quote_garantia_extendida,
        #     "\nSeguro Robo Autopartes: ", quote_seguro_autoparte,
        #     "\nSeguro GAP Porcentaje: ", quote_seguro_gap_porcentaje,
        #     "\nSeguro GAP Monto: ", quote_seguro_gap_monto,
        #     "\nAccesorios: ", quote_accesorios_financiados,
        #     "\nCoberturas Adicionales: ", quote_coberturas,
        #     "\nMonto a Financiar en Cotización: ", quote_monto_financiar,
        #     "\nPlazo (meses): ", quote_plazo,
        #     "\nTasa: ", quote_tasa,
        #     "\nComisión por Apertura %: ", quote_comision_porcentaje,
        #     "\nComisión por Apertura con IVA: ", quote_comision_monto,
        #     "\nDesembolso Inicial: ", quote_desembolso,
        #     "\nCAT AUTOMOTRIZ: ", quote_cat,
        #     "\nFecha: ", quote_fecha,
        #     "\nDatos No. Pago 0: ", _info_0,
        #     "\nDatos No. Pago 1: ", _info_1,
        #     "\nDatos No. Pago 13: ", _info_13,
        #     "\nDatos No. Pago 25: ", _info_25,
        # )
        # print("\033[1;32m_" * 20, "_\033[0m" * 20)

        # DATOS DE LA COTIZACIÓN, CLIENTE Y SOLICITUD
        sheet_quote = archivo_excel.sheet_by_index(1)
        fields = [
            "var_estatus", "var_usuario", "var_escenario", "var_agencia", "var_tipo_credito", "var_producto_finan",
            "var_tipo_producto", "var_tipo_uso", "var_tipo_persona", "var_marca",
            "var_modelo", "var_version", "var_ano", "var_plan",
            "var_accesorios", "var_monto_accesorios", "var_forma_pago_acce",
            "var_g_ext", "var_monto_g_ext", "var_forma_pago_g_ext",
            "var_plazo", "var_seguro_robo_aut", "var_seguro_robo_aut_cobertura",
            "var_gap", "var_gap_forma_pago", "var_seguro_danos", "var_forma_pago", "var_tipo",
            "var_cp", "var_cobertura", "var_fecha", "var_sexo", "var_UDI",
            "var_recibo1", "var_recibo2", "var_subsecuente", "var_autocompara",
            "var_aseguradora", "var_faker", "lada", "telefono", "celular",
            "compania", "email", "var_tbl_amortiza", "var_crearsolicitud",
            "var_clienteSNTD", "var_sexo_solicitud", "var_fechanac", "var_entidad",
            "var_homoclave", "var_curp", "var_edocivil", "var_r_matrimonial",
            "var_dependientes", "var_nivelestudio", "var_ocupacion"
        ]

        # Iterar sobre las filas del Excel
        for numero_fila in range(1, sheet_quote.nrows):
            fila = {}
            for i, field in enumerate(fields):
                value = sheet_quote.cell_value(numero_fila, i)
                if field in ["var_escenario", "var_ano", "var_plazo"]:
                    try:
                        value = int(value)  # Asegúrate de convertir a entero los campos que lo necesiten
                    except ValueError:
                        pass  # Si no se puede convertir, deja el valor tal como está
                fila[field] = value
            # print(fila)
            try:
                if fila['var_estatus'] != 'EJECUTAR':
                    msj = f"El escenario '{fila['var_escenario']}' no se ejecutó porque está marcado como 'PASAR' o no tiene un estatus asignado."
                    logging.info(msj)
                    pass
                else:
                    self.login(fila, fila_environment)
                    self.cotizador_datos_solicitante(fila)
                    self.cotizador_producto(fila, fila_environment)
                    self.cotizador_plan_financiero(fila)
                    self.cotizador_seguro_de_auto(fila)
                    self.cotizador_detalle_financiamiento(fila)
                    self.cotizador_imprimir(fila)
                    self.cotizador_log(fila)
                    self.solicita_credito(fila)
            except Exception as e:
                error_message = f"Error en el escenario {fila['var_escenario']}: {str(e)}"
                logging.error(error_message)
                print(error_message)
                # traceback.print_exc()

    def login(self, fila, fila_environment):
        try:
            driver = self.driver
            login = LoginPage(driver)
            search = SearchPage(driver)
            quote = QuotePage(driver)

            print("                                                                    ")
            print("\033[1;32m*" * 30 + " ESCENARIO # ", fila["var_escenario"], "*\033[0m" * 30)
            if fila_environment["var_environment"] == "BYD_QA":
                if fila["var_producto_finan"] == "ARRENDAMIENTO" or fila["var_tipo_uso"] == "CHOFER PRVADO" or fila["var_tipo_uso"] == "COMERCIAL" or \
                        fila["var_tipo_producto"] == "SEMINUEVO" or fila["var_tipo_producto"] == "MOTO" or fila["var_marca"] != "BYD":
                    print("La configuración no forma parte del proyecto, revisa los datos e intenta de nuevo")
                    self.driver.quit()
            else:
                pass
            driver.get(fila_environment["var_url"])

            # LOGIN
            if fila_environment["var_environment"] != "BYD_QA" and fila_environment["var_environment"] != "OM2":
                login.enter_username(fila["var_usuario"])
                login.enter_password("S1santan")
                login.click_login()
                print("\033[1;32m-" * 1 + " LOGIN EXITOSO " + "-\033[0m" * 1)

                # Búsqueda de Cotizaciones
                search.click_nvaCot()
                print("\033[1;32m-" * 1 + " BÚSQUEDA EXITOSA " + "-\033[0m" * 1)

                # NVA COTIZACIÓN
                time.sleep(2)
                quote.click_agencia()
                quote.enter_agencia(fila["var_agencia"])

            else:
                pass
        except TimeoutException as e:
            error_message = f"Error en el escenario {fila['var_escenario']}: {str(e)}"
            logging.error(error_message)
            traceback.print_exc()

    def cotizador_producto(self, fila, fila_environment):
        cotizador_producto(self.driver, fila, fila_environment)

    def cotizador_plan_financiero(self, fila):
        cotizador_plan_financiero(self, self.driver, fila)

    def cotizador_seguro_de_auto(self, fila):
        cotizador_seguro_de_auto(self.driver, fila)

    def cotizador_detalle_financiamiento(self, fila):
        cotizador_detalle_financiamiento(self, self.driver, fila)

    def cotizador_datos_solicitante(self, fila):
        cotizador_datos_solicitante(self, self.driver, fila)

    def cotizador_imprimir(self, fila):
        cotizador_imprimir(self.driver, fila)

    def cotizador_log(self, fila):
        try:
            print("\033[1;32m-" * 1 + " DATOS DEL SOLICITANTE EXITOSO " + "-\033[0m" * 1)
            comments_all = "GAP: {}, SRA: {}, GE: {}".format(self.commentgap, self.commentrsa, self.commentge)
            # Guarda información en el log
            logging.info("*" * 45 + " ESCENARIO: %s " + "*" * 45 + "\n"
                                                                   "- [Nombre Cliente]: %s\n"
                                                                   "- [PRODUCTO]: %s\n"
                                                                   "- [Vehículo]: %s\n"
                                                                   "- [Valor del Vehículo]: %s\n"
                                                                   "- [Accesorios Vehículo]: %s\n"
                                                                   "- [Coberturas Adicionales]: %s\n"
                                                                   "- [Garantía Extendida]: %s\n"
                                                                   "- [Seguro de Robo de Autopartes]: %s\n"
                                                                   "- [Seguro GAP]: %s\n"
                                                                   "- [Tasa Fija]: %s\n"
                                                                   "- [Plazo]: %s\n"
                                                                   "- [Seguro de Vida y Desempleo]: %s\n"
                                                                   "- [Seguro del vehículo Año 1]: %s\n"
                                                                   "- [Subsecuente]: %s\n"
                                                                   "- [Enganche]: %s\n"
                                                                   f"- [{self.ballon_and_tcm}] %s\n"
                                                                   "- [Tasa Fija Ciclo 2]: %s\n"
                                                                   "- [Plazo Ciclo 2]: %s\n"
                                                                   "- [Comisión Apertura]: %s\n"
                                                                   "- [Desembolso Inicial]: %s\n"
                                                                   "- [Monto a Financiar]: %s\n"
                                                                   "- [Primer Pago]: %s\n"
                                                                   "- [CAT]: %s\n"
                                                                   "- [Cuotas All]: %s\n"
                                                                   "- [Tabla de Amortización]: %s\n"
                                                                   f"- [{self.TblamortizationTCM}] %s\n"
                                                                   "- [Comentarios generales]: %s\n",
                         fila["var_escenario"], self.namecomplete, self.productall, self.lblcar, self.lblpricelist, self.tdaccessoriesamount, self.tdadditionalcoverageamount,
                         self.tdextendedwarrantyamount, self.tdautopartstheftinsuranceamount, self.tdInsuranceGAPAmount, self.tdTasa, self.tdPLazo, self.lblSeguroVd, self.tdSeguro,
                         self.replace_divInsuranceRecipes, self.lblEnganche, '', self.tasaC2, self.plazoC2, self.tdComision, self.tdDesembolso, self.tdFinanciar, self.tdMensualidad, self.cat,
                         self.cuotasall, self.Tblamortization, '', comments_all)

            print("\033[1;32m-" * 1 + " FINISHED COTIZACIÓN" + "-\033[0m" * 1)
            time.sleep(10)
        except TimeoutException as e:
            error_message = f"Error en el escenario {fila['var_escenario']}: {str(e)}"
            logging.error(error_message)
            traceback.print_exc()

    def solicita_credito(self, fila):
        solicita_credito(self.driver, fila)

    def solicitud_datos_personales(self, fila):
        solicitud_datos_personales(self.driver, fila)

    def tearDown(self):
        try:
            time.sleep(1)
            self.driver.quit()
            time.sleep(2)
        except Exception as e:
            error_message = f"Error en el escenario : {str(e)}"
            logging.error(error_message + "validando en el log")


if __name__ == '__main__':
    unittest.main()
